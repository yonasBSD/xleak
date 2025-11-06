#!/usr/bin/env python3
"""
Generate sample Excel files for testing xleak
Run: python3 generate_test_data.py
"""

try:
    from openpyxl import Workbook
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    exit(1)

def create_sample_file():
    wb = Workbook()
    
    # Sheet 1: Sales Data (30 rows x 15 columns)
    ws1 = wb.active
    ws1.title = "Sales"
    
    headers = ["Product", "Jan", "Feb", "Mar", "Apr", "May", "Jun", 
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total", "Avg"]
    ws1.append(headers)
    
    products = ["Widgets", "Gadgets", "Doohickeys", "Thingamajigs", "Sprockets",
                "Flanges", "Cogs", "Bearings", "Gizmos", "Contraptions"]
    
    for i in range(30):
        product = products[i % len(products)]
        if i >= len(products):
            product = f"{product} {i // len(products)}"
        
        monthly_sales = [round(500 + (i * 37.5) + (j * 25.3), 2) for j in range(12)]
        total = round(sum(monthly_sales), 2)
        avg = round(total / 12, 2)
        
        ws1.append([product] + monthly_sales + [total, avg])
    
    # Sheet 2: Employee List
    ws2 = wb.create_sheet("Employees")
    ws2.append(["Name", "Department", "Salary", "Active", "Start Date"])
    ws2.append(["Alice Johnson", "Engineering", 95000, True, "2020-01-15"])
    ws2.append(["Bob Smith", "Sales", 75000, True, "2019-06-01"])
    ws2.append(["Carol White", "Marketing", 68000, True, "2021-03-10"])
    ws2.append(["David Brown", "Engineering", 105000, True, "2018-09-20"])
    ws2.append(["Eve Davis", "Sales", 72000, False, "2022-02-14"])
    
    # Sheet 3: Metrics
    ws3 = wb.create_sheet("Metrics")
    ws3.append(["Metric", "Value", "Change", "Status"])
    ws3.append(["Revenue", 1250000, 12.5, "Up"])
    ws3.append(["Customers", 4580, -2.3, "Down"])
    ws3.append(["Satisfaction", 4.7, 0.3, "Up"])
    ws3.append(["Churn Rate", 3.2, -0.5, "Down"])
    
    wb.save("test_data.xlsx")
    print("✅ Created test_data.xlsx with 3 sheets (Sales: 30 rows x 15 columns)")

if __name__ == "__main__":
    create_sample_file()
