#!/usr/bin/env python3
"""Generate a comprehensive test Excel file with various data types."""

import openpyxl
from datetime import datetime, timedelta
import random

# Create a new workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet 1: Customer Orders - realistic business data
orders = wb.create_sheet("Customer Orders")
orders.append([
    "Order ID", "Customer Name", "Email", "Order Date", "Ship Date",
    "Amount", "Status", "Priority", "Notes"
])

customers = [
    ("Alice Johnson", "alice.j@example.com"),
    ("Bob Smith", "bob.smith@example.com"),
    ("Carol Martinez", "c.martinez@example.com"),
    ("David Chen", "david.chen@example.com"),
    ("Emma Wilson", "emma.w@example.com"),
    ("Frank Brown", "frank.brown@example.com"),
    ("Grace Lee", "grace.lee@example.com"),
    ("Henry Taylor", "h.taylor@example.com"),
]

statuses = ["Pending", "Shipped", "Delivered", "Cancelled", "Processing"]
priorities = ["Low", "Normal", "High", "Urgent"]
notes = [
    "Customer requested gift wrap",
    "Deliver to back door",
    "Call before delivery",
    "Leave with receptionist",
    "Fragile - handle with care",
    "Express shipping",
    "Hold for pickup",
    "",
    "",
]

base_date = datetime(2024, 1, 1)
for i in range(50):
    customer_name, email = random.choice(customers)
    order_date = base_date + timedelta(days=random.randint(0, 180))
    ship_date = order_date + timedelta(days=random.randint(1, 7))
    amount = random.randint(100, 50000) / 10.0
    status = random.choice(statuses)
    priority = random.choice(priorities)
    note = random.choice(notes)

    orders.append([
        f"ORD-{1000 + i}",
        customer_name,
        email,
        order_date,
        ship_date if status != "Cancelled" else None,
        amount,
        status,
        priority,
        note
    ])

# Sheet 2: Project Timeline
projects = wb.create_sheet("Project Timeline")
projects.append([
    "Project", "Team", "Start Date", "End Date", "Budget",
    "Spent", "Complete", "On Track", "Manager"
])

project_names = [
    "Website Redesign", "Mobile App Development", "Database Migration",
    "Security Audit", "Cloud Infrastructure", "API Integration",
    "Customer Portal", "Analytics Dashboard", "Payment Gateway",
    "Inventory System", "CRM Implementation", "Marketing Campaign"
]

teams = ["Engineering", "Design", "DevOps", "Data", "Product", "Marketing"]
managers = ["Sarah Connor", "John Wick", "Diana Prince", "Bruce Wayne", "Tony Stark"]

for i, project in enumerate(project_names):
    start = datetime(2024, 1, 1) + timedelta(days=i * 15)
    duration = random.randint(30, 180)
    end = start + timedelta(days=duration)
    budget = random.randint(50000, 500000)
    spent = int(budget * random.uniform(0.3, 0.9))
    complete = random.randint(25, 95)
    on_track = complete >= 50 and spent < budget * 0.8

    projects.append([
        project,
        random.choice(teams),
        start,
        end,
        budget,
        spent,
        f"{complete}%",
        on_track,
        random.choice(managers)
    ])

# Sheet 3: Product Inventory with long descriptions
inventory = wb.create_sheet("Product Inventory")
inventory.append([
    "SKU", "Product Name", "Description", "Category", "Supplier",
    "Cost", "Price", "Stock", "Reorder", "Last Updated"
])

products = [
    ("Wireless Mouse", "Ergonomic wireless mouse with 2.4GHz connectivity and 6-month battery life", "Electronics", 125.50, 249.99),
    ("Mechanical Keyboard", "RGB backlit mechanical keyboard with Cherry MX switches and aluminum frame", "Electronics", 180.00, 349.99),
    ("USB-C Hub", "7-in-1 USB-C hub with HDMI, ethernet, and SD card reader for laptops", "Accessories", 45.00, 89.99),
    ("Laptop Stand", "Adjustable aluminum laptop stand with cooling vents and cable management", "Accessories", 35.00, 69.99),
    ("Webcam HD", "1080p HD webcam with built-in microphone and automatic light correction", "Electronics", 75.00, 149.99),
    ("Phone Charger", "Fast charging USB-C charger 65W with foldable plug and cable included", "Accessories", 25.00, 49.99),
    ("Screen Protector", "Tempered glass screen protector with anti-fingerprint coating", "Accessories", 8.00, 19.99),
    ("Cable Organizer", "Silicone cable management clips for desk organization and cable routing", "Accessories", 5.00, 12.99),
    ("Desk Lamp", "LED desk lamp with adjustable brightness and color temperature control", "Office", 55.00, 109.99),
    ("Monitor Arm", "Dual monitor arm with gas spring for easy height and angle adjustment", "Office", 95.00, 189.99),
]

suppliers = ["TechCorp", "GlobalSupply", "DirectImport", "MegaWholesale", "PremiumGoods"]

for i, (name, desc, category, cost, price) in enumerate(products):
    stock = random.randint(0, 500)
    reorder = stock < 50
    last_update = datetime.now() - timedelta(days=random.randint(1, 30))

    inventory.append([
        f"SKU-{2000 + i}",
        name,
        desc,
        category,
        random.choice(suppliers),
        cost,
        price,
        stock,
        reorder,
        last_update
    ])

# Sheet 4: Employee Records with dates
employees = wb.create_sheet("Employees")
employees.append([
    "Employee ID", "Full Name", "Department", "Position", "Hire Date",
    "Salary", "Active", "Email", "Location"
])

first_names = ["Alice", "Bob", "Carol", "David", "Emma", "Frank", "Grace", "Henry", "Ivy", "Jack"]
last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Rodriguez", "Martinez"]
departments = ["Engineering", "Sales", "Marketing", "HR", "Finance", "Operations", "Support"]
positions = {
    "Engineering": ["Senior Engineer", "Engineer", "Junior Engineer", "Tech Lead"],
    "Sales": ["Sales Manager", "Account Executive", "Sales Rep"],
    "Marketing": ["Marketing Manager", "Content Writer", "Designer"],
    "HR": ["HR Manager", "Recruiter", "HR Coordinator"],
    "Finance": ["Financial Analyst", "Accountant", "Controller"],
    "Operations": ["Operations Manager", "Coordinator", "Specialist"],
    "Support": ["Support Manager", "Support Agent", "Technical Support"]
}
locations = ["New York", "San Francisco", "Austin", "Seattle", "Boston", "Chicago", "Remote"]

for i in range(30):
    first = random.choice(first_names)
    last = random.choice(last_names)
    dept = random.choice(departments)
    position = random.choice(positions[dept])
    hire_date = datetime(2020, 1, 1) + timedelta(days=random.randint(0, 1500))

    # Salary based on position
    base_salary = 60000
    if "Senior" in position or "Manager" in position or "Lead" in position:
        base_salary = 120000
    elif "Junior" in position or "Coordinator" in position:
        base_salary = 50000

    salary = base_salary + random.randint(-10000, 20000)
    active = random.choice([True, True, True, True, False])  # 80% active

    employees.append([
        f"EMP{1000 + i}",
        f"{first} {last}",
        dept,
        position,
        hire_date,
        salary,
        active,
        f"{first.lower()}.{last.lower()}@company.com",
        random.choice(locations)
    ])

# Format date columns
for sheet in wb.worksheets:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'YYYY-MM-DD'

# Save the file
wb.save("comprehensive_test.xlsx")
print("Created comprehensive_test.xlsx with 4 sheets:")
print(f"  - Customer Orders: {orders.max_row - 1} rows")
print(f"  - Project Timeline: {projects.max_row - 1} rows")
print(f"  - Product Inventory: {inventory.max_row - 1} rows")
print(f"  - Employees: {employees.max_row - 1} rows")
print("\nTest with: cargo run --release -- comprehensive_test.xlsx -i")
