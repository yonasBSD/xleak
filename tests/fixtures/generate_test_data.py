#!/usr/bin/env python3
"""
Generate sample Excel files for testing xleak
Run: python3 generate_test_data.py
"""

try:
    from openpyxl import Workbook
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    exit(1)

def create_sample_file():
    wb = Workbook()
    
    # Sheet 1: Sales Data (30 rows x 15 columns)
    ws1 = wb.active
    ws1.title = "Sales"
    
    headers = ["Product", "Jan", "Feb", "Mar", "Apr", "May", "Jun", 
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total", "Avg"]
    ws1.append(headers)
    
    products = ["Widgets", "Gadgets", "Doohickeys", "Thingamajigs", "Sprockets",
                "Flanges", "Cogs", "Bearings", "Gizmos", "Contraptions"]
    
    for i in range(30):
        product = products[i % len(products)]
        if i >= len(products):
            product = f"{product} {i // len(products)}"

        monthly_sales = [round(500 + (i * 37.5) + (j * 25.3), 2) for j in range(12)]

        # Calculate totals for the row
        total = round(sum(monthly_sales), 2)
        avg = round(total / 12, 2)

        row_data = [product] + monthly_sales + [total, avg]
        ws1.append(row_data)

        # Get the actual row number that was just appended
        actual_row = ws1.max_row

        # Add formulas (openpyxl will keep both the value and formula)
        total_cell = ws1[f'N{actual_row}']
        total_cell.value = f'=SUM(B{actual_row}:M{actual_row})'

        avg_cell = ws1[f'O{actual_row}']
        avg_cell.value = f'=AVERAGE(B{actual_row}:M{actual_row})'
    
    # Sheet 2: Employee List
    ws2 = wb.create_sheet("Employees")
    ws2.append(["Name", "Department", "Salary", "Active", "Start Date", "Bonus (10%)"])

    # Employee data with calculated bonuses
    employees = [
        ["Alice Johnson", "Engineering", 95000, True, "2020-01-15"],
        ["Bob Smith", "Sales", 75000, True, "2019-06-01"],
        ["Carol White", "Marketing", 68000, True, "2021-03-10"],
        ["David Brown", "Engineering", 105000, True, "2018-09-20"],
        ["Eve Davis", "Sales", 72000, False, "2022-02-14"],
    ]

    # Track the actual Excel row number (header is row 1, data starts at row 2)
    for idx, emp in enumerate(employees, start=2):
        name, dept, salary, active, start_date = emp
        # Calculate bonus: 10% if active, 0 otherwise
        bonus = salary * 0.1 if active else 0

        # Append row with calculated bonus
        ws2.append([name, dept, salary, active, start_date, bonus])

        # The row was just appended, so it's at the current max_row
        actual_row = ws2.max_row

        # Add formula to the bonus cell at the actual row
        bonus_cell = ws2[f'F{actual_row}']
        bonus_cell.value = f'=IF(D{actual_row}, C{actual_row}*0.1, 0)'
    
    # Sheet 3: Metrics
    ws3 = wb.create_sheet("Metrics")
    ws3.append(["Metric", "Value", "Change", "Status"])
    ws3.append(["Revenue", 1250000, 12.5, "Up"])
    ws3.append(["Customers", 4580, -2.3, "Down"])
    ws3.append(["Satisfaction", 4.7, 0.3, "Up"])
    ws3.append(["Churn Rate", 3.2, -0.5, "Down"])
    
    wb.save("test_data.xlsx")
    print("✅ Created test_data.xlsx with 3 sheets")
    print("   - Sales: 30 rows with SUM and AVERAGE formulas")
    print("   - Employees: 5 rows with IF formulas for bonuses")
    print("   - Metrics: 4 rows")
    print()
    print("⚠️  IMPORTANT: Open test_data.xlsx in Excel and save it to calculate formulas.")
    print("   (openpyxl doesn't evaluate formulas - Excel needs to cache the results)")

if __name__ == "__main__":
    create_sample_file()
