#!/usr/bin/env python3
"""
Generate Excel file with Excel Table structures for testing (issue #18).
Run: python3 generate_test_tables.py
"""

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import random
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    exit(1)

def create_products_table(wb):
    """Create sheet with Products table."""
    ws = wb.create_sheet("ProductsTable")

    # Add data
    data = [
        ["ProductID", "ProductName", "Category", "Price", "Stock", "Supplier"],
        [1001, "Wireless Mouse", "Electronics", 24.99, 150, "TechCorp"],
        [1002, "USB-C Cable", "Accessories", 9.99, 500, "CableWorld"],
        [1003, "Laptop Stand", "Accessories", 34.99, 75, "OfficePlus"],
        [1004, "Mechanical Keyboard", "Electronics", 89.99, 45, "KeyMasters"],
        [1005, "Webcam HD", "Electronics", 59.99, 120, "TechCorp"],
        [1006, "Phone Charger", "Accessories", 19.99, 300, "ChargeIt"],
        [1007, "Monitor Arm", "Office", 79.99, 30, "OfficePlus"],
        [1008, "Desk Lamp LED", "Office", 44.99, 85, "LightWorks"],
        [1009, "Cable Organizer", "Accessories", 12.99, 200, "OfficePlus"],
        [1010, "USB Hub", "Electronics", 29.99, 160, "TechCorp"],
    ]

    for row in data:
        ws.append(row)

    # Create table
    tab = Table(displayName="Products", ref="A1:F11")

    # Add a table style
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(tab)

    return ws

def create_sales_table(wb):
    """Create sheet with Sales table including calculated column."""
    ws = wb.create_sheet("SalesTable")

    # Add data
    data = [
        ["OrderID", "ProductID", "Quantity", "UnitPrice", "Total", "OrderDate"],
        [5001, 1001, 2, 24.99, 49.98, "2024-01-15"],
        [5002, 1003, 1, 34.99, 34.99, "2024-01-16"],
        [5003, 1002, 5, 9.99, 49.95, "2024-01-16"],
        [5004, 1004, 1, 89.99, 89.99, "2024-01-17"],
        [5005, 1005, 2, 59.99, 119.98, "2024-01-18"],
        [5006, 1001, 3, 24.99, 74.97, "2024-01-19"],
        [5007, 1006, 4, 19.99, 79.96, "2024-01-20"],
        [5008, 1008, 1, 44.99, 44.99, "2024-01-21"],
        [5009, 1002, 10, 9.99, 99.90, "2024-01-22"],
        [5010, 1007, 2, 79.99, 159.98, "2024-01-23"],
    ]

    for row in data:
        ws.append(row)

    # Create table
    tab = Table(displayName="Sales", ref="A1:F11")

    # Add a table style
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(tab)

    return ws

def create_employees_table(wb):
    """Create sheet with Employees table."""
    ws = wb.create_sheet("EmployeesTable")

    # Add data
    first_names = ["Alice", "Bob", "Carol", "David", "Emma", "Frank", "Grace", "Henry"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis"]
    departments = ["Engineering", "Sales", "Marketing", "HR", "Finance"]

    data = [["EmpID", "FirstName", "LastName", "Department", "Salary", "HireDate"]]

    for i in range(1, 16):
        data.append([
            2000 + i,
            random.choice(first_names),
            random.choice(last_names),
            random.choice(departments),
            random.randint(50000, 120000),
            f"20{random.randint(18,24)}-{random.randint(1,12):02d}-{random.randint(1,28):02d}"
        ])

    for row in data:
        ws.append(row)

    # Create table
    tab = Table(displayName="Employees", ref="A1:F16")

    # Add a table style
    style = TableStyleInfo(
        name="TableStyleLight11",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(tab)

    return ws

def create_unformatted_data(wb):
    """Create sheet with regular range (non-table) for comparison."""
    ws = wb.create_sheet("UnformattedData")

    # Add regular data without table formatting
    data = [
        ["Region", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Q4 Sales"],
        ["North", 125000, 132000, 145000, 158000],
        ["South", 98000, 105000, 112000, 121000],
        ["East", 156000, 162000, 175000, 188000],
        ["West", 187000, 195000, 208000, 221000],
    ]

    for row in data:
        ws.append(row)

    return ws

def main():
    """Generate the Excel tables test file."""
    print("Generating test_tables.xlsx...")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets with tables
    print("  Creating ProductsTable...")
    create_products_table(wb)

    print("  Creating SalesTable...")
    create_sales_table(wb)

    print("  Creating EmployeesTable...")
    create_employees_table(wb)

    print("  Creating UnformattedData (comparison)...")
    create_unformatted_data(wb)

    # Save file
    filename = "test_tables.xlsx"
    wb.save(filename)

    print(f"\n✓ Created {filename}")
    print("\nSheets:")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        tables = ws.tables
        if tables:
            table_names = [t.displayName for t in tables.values()]
            print(f"  - {sheet}: {ws.max_row - 1} data rows, table: {', '.join(table_names)}")
        else:
            print(f"  - {sheet}: {ws.max_row - 1} data rows (no table)")

    print("\nTest with:")
    print(f"  ./target/release/xleak {filename} -i")
    print(f"  ./target/release/xleak {filename} --list-tables")
    print(f"  ./target/release/xleak {filename} --table Products")

if __name__ == "__main__":
    main()
