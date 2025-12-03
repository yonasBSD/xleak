#!/usr/bin/env python3
"""Generate comprehensive test Excel file with all data types and edge cases."""

import openpyxl
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter

def create_data_types_sheet(wb):
    """Create sheet with all data types."""
    ws = wb.create_sheet("DataTypes")

    # Headers
    ws.append(["Type", "Value", "Description"])

    # Strings
    ws.append(["String", "Hello World", "Simple string"])
    ws.append(["String", "Text with spaces   ", "String with trailing spaces"])
    ws.append(["String", "123456", "Number as string"])
    ws.append(["String", "=TEXT", "String that looks like formula"])

    # Numbers
    ws.append(["Integer", 42, "Positive integer"])
    ws.append(["Integer", -273, "Negative integer"])
    ws.append(["Float", 3.14159, "Decimal number"])
    ws.append(["Float", 2.5e10, "Scientific notation"])
    ws.append(["Percentage", 0.75, "Percentage (75%)"])
    ws['B9'].number_format = '0%'

    # Booleans
    ws.append(["Boolean", True, "TRUE value"])
    ws.append(["Boolean", False, "FALSE value"])

    # Empty
    ws.append(["Empty", None, "Empty cell"])

    # Currency
    ws.append(["Currency", 1234.56, "US Dollar"])
    ws['B13'].number_format = '$#,##0.00'
    ws.append(["Currency", 9876.54, "Euro"])
    ws['B14'].number_format = '€#,##0.00'

    return ws

def create_formulas_sheet(wb):
    """Create sheet with various formulas."""
    ws = wb.create_sheet("Formulas")

    # Headers
    ws.append(["Formula Type", "Formula", "Result", "Description"])

    # Sample data
    ws.append(["Data", 10, None, "Sample values"])
    ws.append(["Data", 20, None, None])
    ws.append(["Data", 30, None, None])
    ws.append(["Data", 40, None, None])
    ws.append(["Data", 50, None, None])

    # Arithmetic formulas
    ws.append(["SUM", "=SUM(B2:B6)", None, "Sum of range"])
    ws.append(["AVERAGE", "=AVERAGE(B2:B6)", None, "Average of range"])
    ws.append(["MIN", "=MIN(B2:B6)", None, "Minimum value"])
    ws.append(["MAX", "=MAX(B2:B6)", None, "Maximum value"])
    ws.append(["COUNT", "=COUNT(B2:B6)", None, "Count numbers"])

    # Logic formulas
    ws.append(["IF", "=IF(B2>15,\"High\",\"Low\")", None, "Conditional"])
    ws.append(["AND", "=AND(B2>5,B2<15)", None, "Logical AND"])
    ws.append(["OR", "=OR(B2>100,B2<5)", None, "Logical OR"])

    # Math formulas
    ws.append(["ROUND", "=ROUND(3.14159,2)", None, "Round to 2 decimals"])
    ws.append(["ABS", "=ABS(-42)", None, "Absolute value"])
    ws.append(["SQRT", "=SQRT(144)", None, "Square root"])

    return ws

def create_internationalization_sheet(wb):
    """Create sheet with international characters (issue #11)."""
    ws = wb.create_sheet("Internationalization")

    # Headers
    ws.append(["Language", "Text", "Characters", "Description"])

    # German
    ws.append(["German", "Größe", "ö, ß", "Umlauts and eszett"])
    ws.append(["German", "Äpfel und Übung", "Ä, Ü", "More umlauts"])
    ws.append(["German", "§123 Paragraph", "§", "Section sign"])

    # Turkish
    ws.append(["Turkish", "İstanbul", "İ", "Dotted capital I"])
    ws.append(["Turkish", "ışık", "ı, ş", "Dotless i, cedilla"])
    ws.append(["Turkish", "çalışma", "ç, ş", "Turkish characters"])

    # Chinese (Simplified)
    ws.append(["Chinese", "简体中文", "简体中文", "Simplified Chinese"])
    ws.append(["Chinese", "表格文件", "表格文件", "Table file"])
    ws.append(["Chinese", "测试数据", "测试数据", "Test data"])

    # Japanese
    ws.append(["Japanese", "日本語", "日本語", "Japanese language"])
    ws.append(["Japanese", "テスト", "テスト", "Test (katakana)"])

    # Mixed
    ws.append(["Mixed", "Café résumé", "é", "French accents"])
    ws.append(["Mixed", "Москва", "Cyrillic", "Russian (Moscow)"])
    ws.append(["Mixed", "naïve", "ï", "Diaeresis"])

    return ws

def create_multiline_cells_sheet(wb):
    """Create sheet with multi-line cells (issue #16)."""
    ws = wb.create_sheet("MultilineCells")

    # Headers
    ws.append(["Line Count", "Content", "Description"])

    # 1 line
    ws.append([1, "Single line", "No newlines"])

    # 5 lines
    lines_5 = "\n".join([f"Line {i+1}" for i in range(5)])
    ws.append([5, lines_5, "Short multi-line"])

    # 10 lines
    lines_10 = "\n".join([f"Line {i+1}" for i in range(10)])
    ws.append([10, lines_10, "Medium multi-line"])

    # 20 lines (original issue #16 case)
    lines_20 = "\n".join([f"Line {i+1}" for i in range(20)])
    ws.append([20, lines_20, "Issue #16 test case"])

    # 50 lines
    lines_50 = "\n".join([f"Line {i+1}" for i in range(50)])
    ws.append([50, lines_50, "Large multi-line"])

    # 100 lines
    lines_100 = "\n".join([f"Line {i+1}" for i in range(100)])
    ws.append([100, lines_100, "Very large multi-line"])

    # Multi-line with international characters
    multi_intl = "Line 1: English\nLine 2: Deutsch (ä, ö, ü)\nLine 3: 中文\nLine 4: Türkçe"
    ws.append([4, multi_intl, "Mixed languages"])

    # Multi-line with empty lines
    with_empty = "Line 1\n\nLine 3 (empty line 2)\n\n\nLine 6 (empty 4-5)"
    ws.append([6, with_empty, "Contains empty lines"])

    return ws

def create_date_edge_cases_sheet(wb):
    """Create sheet with date edge cases (issue #25)."""
    ws = wb.create_sheet("DateEdgeCases")

    # Headers
    ws.append(["Description", "Date", "Serial", "Notes"])

    # 1900 era dates (leap year bug zone)
    ws.append(["First day", datetime(1900, 1, 1), 1, "Excel serial 1"])
    ws.append(["Late January", datetime(1900, 1, 31), 31, "Before leap bug"])
    ws.append(["Day before fake leap", datetime(1900, 2, 28), 59, "Serial 59"])
    # Note: Excel allows 1900-02-29 (serial 60) but Python doesn't, so we skip it
    ws.append(["Day after fake leap", datetime(1900, 3, 1), 61, "Serial 61, leap bug applies"])

    # Modern dates
    ws.append(["Y2K", datetime(2000, 1, 1), 36526, "Year 2000"])
    ws.append(["Real leap day", datetime(2000, 2, 29), 36585, "2000 was a leap year"])
    ws.append(["Recent date", datetime(2024, 1, 1), 45292, "2024 start"])
    ws.append(["Today", datetime(2024, 12, 3), 45629, "Current date"])

    # Issue #25 specific dates
    ws.append(["Issue #25 test", datetime(2025, 11, 19, 11, 18, 20), None, "Date from screenshot"])
    ws.append(["Issue #25 test 2", datetime(2025, 11, 19, 11, 18, 22), None, "Second row from screenshot"])

    # Date with time
    ws.append(["Date + Time", datetime(2024, 6, 15, 14, 30, 0), None, "With time component"])
    ws.append(["Midnight", datetime(2024, 6, 15, 0, 0, 0), None, "Time = 00:00:00"])
    ws.append(["Just before midnight", datetime(2024, 6, 15, 23, 59, 59), None, "Time = 23:59:59"])

    # Format serial column as numbers
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 3).value is not None:
            ws.cell(row, 3).number_format = '0'

    return ws

def create_edge_cases_sheet(wb):
    """Create sheet with various edge cases."""
    ws = wb.create_sheet("EdgeCases")

    # Headers
    ws.append(["Type", "Value", "Description"])

    # Very long string
    long_str = "A" * 1000
    ws.append(["Long String", long_str, "1000 characters"])

    # Very long string with spaces
    long_with_spaces = " ".join(["Word"] * 200)
    ws.append(["Long w/ Spaces", long_with_spaces, "200 words"])

    # Whitespace variations
    ws.append(["Leading Space", "   Leading", "3 leading spaces"])
    ws.append(["Trailing Space", "Trailing   ", "3 trailing spaces"])
    ws.append(["Tab Character", "Before\tAfter", "Contains tab"])
    ws.append(["Newline", "Line1\nLine2", "Contains newline"])
    ws.append(["Multiple Spaces", "Word    Word", "Multiple internal spaces"])

    # Number edge cases
    ws.append(["Large Number", 9999999999999, "Large integer"])
    ws.append(["Small Decimal", 0.000000001, "Very small decimal"])
    ws.append(["Zero", 0, "Zero value"])
    ws.append(["Negative Zero", -0.0, "Negative zero"])

    # Special strings
    ws.append(["Quotes", 'He said "Hello"', "Contains quotes"])
    ws.append(["Apostrophe", "Don't", "Contains apostrophe"])
    ws.append(["Comma", "Last, First", "Contains comma"])
    ws.append(["Semicolon", "A;B;C", "Contains semicolons"])

    # Mixed content
    ws.append(["Number String", "  42  ", "Number with spaces"])
    ws.append(["Decimal String", "3.14", "Decimal as string"])
    ws.append(["Date String", "2024-01-01", "Date as string"])

    return ws

def main():
    """Generate the comprehensive test file."""
    print("Generating test_comprehensive.xlsx...")

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create all sheets
    print("  Creating DataTypes sheet...")
    create_data_types_sheet(wb)

    print("  Creating Formulas sheet...")
    create_formulas_sheet(wb)

    print("  Creating Internationalization sheet...")
    create_internationalization_sheet(wb)

    print("  Creating MultilineCells sheet...")
    create_multiline_cells_sheet(wb)

    print("  Creating DateEdgeCases sheet...")
    create_date_edge_cases_sheet(wb)

    print("  Creating EdgeCases sheet...")
    create_edge_cases_sheet(wb)

    # Save file
    output_path = "test_comprehensive.xlsx"
    wb.save(output_path)

    print(f"\n✓ Created {output_path}")
    print("\nSheets:")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"  - {sheet}: {ws.max_row - 1} data rows")

    print("\n⚠️  IMPORTANT: Open this file in Excel and save it to cache formula results!")
    print("   Formulas will show as #NAME? until Excel calculates them.")

if __name__ == "__main__":
    main()
